"""
Export utility functions for MailTask application.
Handles Excel export functionality for customers and tasks.
"""
import json
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_customers_to_excel(rows):
    """
    Export customers data to Excel format.
    
    Args:
        rows: List of customer rows from database (sqlite3.Row objects)
    
    Returns:
        tuple: (BytesIO object, filename) - Excel file data and filename
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['Name', 'Email Suffix', 'Country', 'Website', 'Remark', 'Company Name', 'Tel', 'Source', 'Address', 'Business Type', 'Created At', 'Created By']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row in rows:
        ws.append([
            row['name'] or '',
            row['email_suffix'] or '',
            row['country'] or '',
            row['website'] or '',
            (row['remark'] or '').replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n'),
            row['company_name'] or '',
            row['tel'] or '',
            row['source'] or '',
            row['address'] or '',
            row['business_type'] or '',
            row['created_at'] or '',
            row['created_by'] or ''
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return output, filename


def export_tasks_to_excel(rows):
    """
    Export tasks data to Excel format.
    
    Args:
        rows: List of task rows from database (sqlite3.Row objects)
    
    Returns:
        tuple: (BytesIO object, filename) - Excel file data and filename
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Headers
    headers = ['Sequence', 'Business Type', 'Customer', 'Company Name', 'Email', 'Type', 'Template', 'Status', 'Deadline', 'Created At', 'Created By', 'Source', 'Last Updated', 'Attachments']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row in rows:
        # Parse attachments to get filenames
        attachments_text = ''
        try:
            if row['attachments']:
                attachments = json.loads(row['attachments'])
                if isinstance(attachments, list):
                    attachments_text = ', '.join([att.get('filename', 'attachment') for att in attachments if isinstance(att, dict)])
        except:
            pass
        
        ws.append([
            row['sequence'] or '',
            row['business_type'] or '',
            row['customer'] or '',
            row['company_name'] or '',
            row['email'] or '',
            row['catalogue'] or '',
            row['template'] or '',
            row.get('status', 'open') or 'open',
            row['deadline'] or '',
            row['created_at'] or '',
            row['created_by'] or '',
            row['source'] or '',
            row['updated_at'] or '',
            attachments_text
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'tasks_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return output, filename

